from openpyxl.reader.excel import load_workbook

from SetUp import SetUp as S1
import pandas as pd
import logging
import os
import sys
import subprocess
import datetime
from pathlib import Path
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.message import EmailMessage
import glob
import shutil
import psutil


InputDir = S1.InputFile
OutputDir = S1.OutFile
ErrorDir = S1.ErrorFile
LogDir = S1.LogFile
ResponseFileDir = S1.ResponseFile
DumpedFileDir = S1.DumpedFile


dir_path = os.getcwd()
dir_path = str(dir_path) + "\\"
LOG_FILE = ""
EMailSubject = ""
EmailBody = ""
Footer = ""
ABORT = False
ProcessTheFile = False
ErrorReason = ""
#FileHeader = "Report_Date,business_date,batch_timestamp,account_uuid,plan_uuid,schedule_id,institution_id,Product_ID,error,error_message,field_path\n"

FORMATTER = logging.Formatter("%(asctime)s — %(levelname)s — %(message)s")
FileName = ""

LogDateTime = datetime.datetime.now()

OutFileName = S1.OutFileNamePrefix + str(LogDateTime.strftime('%Y%m%d%H%M%S'))

ProcessingTime  = str(LogDateTime.strftime("%Y-%m-%d %H:%M:%S"))

term1 = "Start Time"
term2 = "Finish Time"


def get_console_handler():
   console_handler = logging.StreamHandler(sys.stdout)
   console_handler.setFormatter(FORMATTER)
   return console_handler

###################################################################################################


def get_file_handler():
   file_handler = logging.FileHandler(LOG_FILE)
   file_handler.setFormatter(FORMATTER)
   return file_handler

###################################################################################################


def get_logger(logger_name):
   logger = logging.getLogger(logger_name)
   logger.setLevel(logging.INFO)
   logger.addHandler(get_console_handler())
   logger.addHandler(get_file_handler())
   logger.propagate = False

   return logger


###################################################################################################


def DirectoryValidation(InputDir, OutputDir, ErrorDir, LogDir, ResponseFileDir, DumpedFileDir):
    ErrorFlag = True
    ErrorNumber = 0
    Message = ""
    if InputDir == "":
        ErrorFlag = True
        Message = "Environment variable for Input folder is not set"
    elif OutputDir == "":
        ErrorFlag = True
        Message = "Environment variable for Output folder is not set"
    elif ErrorDir == "":
        ErrorFlag = True
        Message = "Environment variable for Error folder is not set"
    elif LogDir == "":
        ErrorFlag = True
        Message = "Environment variable for Log folder is not set"
    elif ResponseFileDir == "":
        ErrorFlag = True
        Message = "Environment variable for ResponseFile folder is not set"
    elif DumpedFileDir == "":
        ErrorFlag = True
        Message = "Environment variable for DumpedFile folder is not set"
    else:
        ErrorFlag = False
        Message = "No Error in setting the folders variables"

    print(Message)

    if ErrorFlag == False:
        InputDir = InputDir + "\\"
        OutputDir = OutputDir + "\\"
        ErrorDir = ErrorDir + "\\"
        LogDir = LogDir + "\\"
        ResponseFileDir = ResponseFileDir + "\\"
        DumpedFileDir = DumpedFileDir + "\\"

    return InputDir, OutputDir, ErrorDir, LogDir, ResponseFileDir, DumpedFileDir, ErrorFlag    

###################################################################################################


def ReadingTheFile(FileName):
    File = InputDir + FileName
    Message = ""
    Count = 0

    with open(File) as f:
        for line in f:
            line.strip().split('/n')
            LineToPrint = line.strip()
            Count += 1

            if LineToPrint != "":
                Message = LineToPrint
                WritingTheFile(Message)


###################################################################################################


def ReadingTheFileDF(FileName, sheets):
    MessageLogger.info("INSIDE ReadingTheFileDF WITH FILE: " + FileName)
    File = InputDir + FileName
    Message = ""
    Count = 0

    DataDict = {}

    for key, values in zip(S1.Columns, S1.DataTypes):
        DataDict[key] = values

    # DF = pd.read_excel(File, sheet_name=S1.SheetName, usecols=S1.Columns)
    DF = pd.read_excel(File, sheet_name=sheets, usecols=S1.Columns)

    MessageLogger.info("DATA CONVERTED FROM XLS TO DF")

    DF1 = pd.DataFrame()
    LoopCount = 0

    for key in S1.Columns:    
        if(DataDict[key] == 'VARCHAR'):
            DF1.insert(LoopCount, key, "'" + DF[key].astype(str) + "'", True)
        elif(DataDict[key] == 'DATETIME'):
            #DF1.insert(LoopCount, key, "'" + (pd.to_datetime(DF[key])).dt.date.astype(str) + "'", True)
            DF1.insert(LoopCount, key, "'" + DF[key].astype(str).str[:10] + "'", True)
        else:
            DF1.insert(LoopCount, key, DF[key].astype(str), True)
        LoopCount+=1

        #to_datetime(df['timestamp_str']).dt.date

    MessageLogger.info("DATA FORMATTED FROM DF TO DF1")

    DF2 = pd.DataFrame()
    LoopCount = 0
    for key in S1.Columns: 
        #print(LoopCount)
        if(LoopCount == 0):
            DF2.insert(LoopCount, 'Value'+str(LoopCount), S1.Prefix+DF1[key], True)
        else:
            DF2.insert(LoopCount, 'Value'+str(LoopCount), DF2['Value'+str(LoopCount-1)]+', '+DF1[key])
        
        LoopCount+=1

    DF2.update(DF2['Value'+str(LoopCount-1)]+S1.Suffix)

    MessageLogger.info("DATA FORMATTED FROM DF1 TO DF2")

    DF3 = DF2['Value'+str(LoopCount-1)]

    MessageLogger.info("DATA FORMATTED FROM DF2 TO DF3")

    MessageLogger.debug(DF3)

    DF3.to_csv('DF.txt', sep=',', index= False, header=False)

    Count = 0
    with open(InputDir + 'DF.txt') as f:
        for line in f:
            line.strip().split('/n')
            LineToPrint = line.strip()
            LineToPrint = LineToPrint.replace('"', '')
            Count += 1

            if LineToPrint != "":
                Message = LineToPrint
                WritingTheFile(Message)
        f.close()

    os.remove(InputDir + 'DF.txt')



###################################################################################################


def WritingTheFile(Message):

    f = open(OutputFile, 'a')

    #OutputMessage = str(LogDateTime.strftime('%Y-%m-%d %H:%M:%S')) + "," + str(LogDateTime.strftime('%Y-%m-%d')) + "," + str(LogDateTime.strftime('%H:%M:%S')) + "," + Message + ",1,6981,7133," + OutFileName + ",NA,NA\n" #production
    # OutputMessage = "INSERT INTO #TempErrorPlans VALUES ('" + Message + "')\n"
    OutputMessage = Message + "\n"	
    # OutputMessage = str(LogDateTime.strftime('%Y-%m-%d %H:%M:%S')) + "," + str(LogDateTime.strftime('%Y-%m-%d')) + "," + str(LogDateTime.strftime('%H:%M:%S')) + "," + Message + ",1,6969,7131,ManualPlanIDCorrection\n" #local

    f.write(OutputMessage)

    f.close()



###################################################################################################


def get_sheet_names(file_path):
    try:
        # MessageLogger.info("file_path: " + file_path)
        workbook = load_workbook(filename=file_path)
        sheet_names = workbook.sheetnames
        return sheet_names
    except Exception as e:
        print(f"Error occurred: {e}")
        return []


def get_active_sheets(file_path):
    active_sheets = []
    try:
        workbook = load_workbook(filename=file_path)
        for sheet in workbook.sheetnames:
            if not workbook[sheet].sheet_state == 'hidden':
                active_sheets.append(sheet)
        return active_sheets
    except Exception as e:
        print(f"Error occurred: {e}")
        return []
    

def close_open_file_handles(file_path):
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # MessageLogger.info(f"proc: {proc}")
            for handle in proc.open_files():
                if handle.path == file_path:
                    MessageLogger.info(f"proc: {proc}")
                    MessageLogger.info(f"handle.path: {handle.path}")
                    MessageLogger.info(f"file_path: {file_path}")
                    proc.terminate()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue


if __name__ == '__main__':

    Error = True

    InputDir, OutputDir, ErrorDir, LogDir, ResponseFileDir, DumpedFileDir, Error = DirectoryValidation(InputDir, OutputDir, ErrorDir, LogDir, ResponseFileDir, DumpedFileDir)

    if Error is False:
        
        LOG_FILE = LogDir + "LOG_" + str(LogDateTime.strftime("%Y%m%d%H%M%S")) + ".log"
        MessageLogger = get_logger(LOG_FILE)

        # OutFileName = OutFileName + ".txt"
        # OutputFile = ResponseFileDir + OutFileName
        # MessageLogger.info("ResponseFile to be processed: " + OutputFile)

        # f = open(OutputFile, 'a')
        # #f.write(FileHeader)
        # f.close()

        Message = "File Processing starts ..."
        MessageLogger.info(Message)
        Message = "START TIME::: " + str(LogDateTime)
        MessageLogger.info(Message)

        try:
            Path = InputDir
            MessageLogger.info(Path)
            os.chdir(Path)

            LoopCount = 0
            JobID = 0

            for File in glob.glob("*"):
                FileName = File
                Message = "********************" + FileName + "*****************************"
                MessageLogger.info(Message)
                LoopCount = LoopCount + 1
                MessageLogger.info(f"LoopCount: {LoopCount}")
                sheet_names = get_active_sheets(FileName)
                MessageLogger.info(sheet_names)
                if S1.ReadAllFiles == 1:
                    for sheets in sheet_names:
                        MessageLogger.info("Processing File: " + FileName)
                        MessageLogger.info("Processing sheet: " + sheets)
                        OutFileName = S1.OutFileNamePrefix #+ str(LogDateTime.strftime('%Y%m%d%H%M%S'))
                        OutFileName = OutFileName + sheets + ".txt"
                        OutputFile = ResponseFileDir + OutFileName
                        MessageLogger.info("ResponseFile to be processed: " + OutputFile)
                        f = open(OutputFile, 'a')
                        f.close()
                        ReadingTheFileDF(FileName, sheets)
                        close_open_file_handles(InputDir + FileName)
                else:
                    if S1.SheetName in sheet_names:
                        MessageLogger.info("Processing File: " + FileName)
                        MessageLogger.info("Processing sheet: " + S1.SheetName)
                        OutFileName = S1.OutFileNamePrefix  # + str(LogDateTime.strftime('%Y%m%d%H%M%S'))
                        OutFileName = OutFileName + S1.SheetName + ".txt"
                        OutputFile = ResponseFileDir + OutFileName
                        MessageLogger.info("ResponseFile to be processed: " + OutputFile)
                        f = open(OutputFile, 'a')
                        f.close()
                        ReadingTheFileDF(FileName, S1.SheetName)
                        close_open_file_handles(InputDir + FileName)
                    else:
                        MessageLogger.info("Please provide correct sheet name")
                shutil.move(InputDir + FileName, OutputDir + FileName)
        except Exception:
            MessageLogger.error("ERROR IN BLOCK", exc_info=True)
            shutil.move(InputDir + FileName, ErrorDir + FileName)

        MessageLogger.info("*********************** FINISHED ***********************")
    else:
        print("ERROR ENCOUNTERED :: " + str(Error))
