import pandas as pd
import requests
import base64
import json
from openpyxl import load_workbook

def read_excel_and_call_jira_api(excel_file_path, jira_url, username, api_token):
    # Read the Excel file into a pandas DataFrame
    try:
        df = pd.read_excel(excel_file_path,sheet_name="Runfor")
    except FileNotFoundError:
        print(f"Error: File '{excel_file_path}' not found.")
        return
    except Exception as e:
        print(f"Error while reading Excel file: {e}")
        return

    # Prepare basic authentication headers
    auth_str = f"{username}:{api_token}"
    base64_auth_str = base64.b64encode(auth_str.encode()).decode()
    headers = {
        "Authorization": f"Basic {base64_auth_str}",
        "Content-Type": "application/json",
    }

    for index, row in df.iterrows():
        CreateTask=row["CreateTask"]
        Logwork=row["Logwork"]

    IssueType_matadata =     {
        "Story": "10001",
        "Task": "10002",
        "Scrum Ceremonies": "10059",
        "Specifications": "10033",
        "Use Case": "10056",
        "CT Query": "10112",
        "Root Cause Analysis": "10109",
        "CARDS Change Request": "10126",
        "Data Analysis": "10131",
        "Release Management": "10130",
        "Release Planning": "10132",
        "Specification Review": "10125",
        "Training": "10110",
        "Analysis": "10065",
        "Meetings/Walkthrough": "10027",
        "Project Management": "10060",
        "Coding": "10025",
        "Design": "10023",
        "Ceremonies": "10038",
        "Unit Test": "10030",
        "Code Review": "10043",
        "Documentation": "10043",
        "Peer Support": "10046",
        "Production Support": "10102",
        "Design Review": "10026",
        "Use Case Creation": "10055",
        "Use Case Review": "10058"
    }
    
    
    error = 0  
    
    # Assuming the Excel file has columns such as 'summary', 'description', 'project', etc.
    if CreateTask ==1:
        try:
            df = pd.read_excel(excel_file_path,sheet_name="CreateTask")
        except FileNotFoundError:
            print(f"Error: File '{excel_file_path}' not found.")
            error = -1 
            return
        except Exception as e:
            print(f"Error while reading Excel file: {e}")
            error = -1 
            return

        for index, row in df.iterrows():
            error = 0  
            # Extract data from Excel row
            id	        = row['id']
            components	= row['components']
            description	= row['description']
            duedate	    = row['duedate']
            startdate	    = row['startdate']
            environment	= row['environment']
            issuetype	= row['issuetype']
            parent	    = row['parent']
            projectid	= row['projectid']
            summary	= row['summary']
            originalEstimate	= row['originalEstimate']
            Logtime	= row['Logtime']
            comment	= row['comment']
            started	= row['started']
            timeSpent= row['timeSpent']
            Completed = row['Completed']
            transitions= row['transitions']
            customfield_10015 = row['startdate']
            transition_done= row['transition_done']
            parent = parent.strip()
            # Add more columns as needed based on your Jira requirements
            invalidissuetype = 0 
            # Create the payload for the Jira API request
            # print(f" before {issuetype}")
            if issuetype in IssueType_matadata:
                print("issuetype exists in the dictionary.")
                issuetype = IssueType_matadata[issuetype]
            else:
                print(f"invalid issuetype = {issuetype}")
                invalidissuetype = 1   
            
            # print(issuetype)
            if Completed == "New":
                if invalidissuetype == 0:
                    payload = {
                    "fields": {
                        "assignee": {
                        "id": id
                        },
                        "components": [
                        {
                            "id": components
                        }
                        ],
                        "description": description,
                        "duedate": duedate,
                        "customfield_10015" : customfield_10015 ,
                        "environment": environment,
                        "issuetype": {
                        "id": issuetype
                        },
                        "labels": [
                        ""
                        ],
                        "parent": {
                        "key": parent
                        },
                        "project": {
                        "id": projectid
                        },
                        "summary": summary,
                        "timetracking": {
                        "originalEstimate": originalEstimate,
                        "remainingEstimate": ""
                        }
                    }
                    
                    }

                    try:
                        # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                        print(payload)
                        response = requests.post(f"{jira_url}/rest/api/2/issue/", json=payload, headers=headers)

                        # Process the response
                        if response.status_code == 201:
                            print(f"Jira issue created successfully for row {index}: {response.json()}")
                            requestresponse  = response.json()
                        else:
                            print(f"Jira issue creation failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                            error = 1
                    except requests.exceptions.RequestException as e:
                        print(f"Error while making Jira API call for row {index}: {e}")
                        error = 2
                    
                    
                    if transitions == 11 and response.status_code == 201:
                        key1= requestresponse["key"]
                        payload_transitions = {
                            "transition": {
                            "id":"11"
                            }
                        }
                        try:
                            # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                            response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/transitions", json=payload_transitions, headers=headers)

                            # Process the response
                            if response.status_code == 204:
                                print(f"transitions sucessfully for row {index}: {response.status_code}")
                                # requestresponse  = response.json()
                            else:
                                print(f"transitions failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                                error = 3
                        except requests.exceptions.RequestException as e:
                            print(f"Error while making transitions Jira API call for row {index}: {e}")   
                            error = 4
                        
                    if Logtime == "Yes" and response.status_code == 204:
                        key1= requestresponse["key"]
                        payloadlogwork = {
                            
                                "comment": comment,
                                "started": started,
                                "timeSpent": timeSpent
                            
                        }
                        try:
                            # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                            response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/worklog", json=payloadlogwork, headers=headers)

                            # Process the response
                            if response.status_code == 201:
                                print(f"timelogged sucessfully for row {index}: {response.json()}")
                                # requestresponse  = response.json()
                            else:
                                print(f"timelog failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                                error = 6
                        except requests.exceptions.RequestException as e:
                            print(f"Error while making timelog Jira API call for row {index}: {e}") 
                            error = 7
                            
                        if transition_done == 21 and  response.status_code == 201:
                            # key1= requestresponse["key"]
                            payload_transitions = {
                                "transition": {
                                "id":"21"
                                }
                            }
                            try:
                                # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                                response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/transitions", json=payload_transitions, headers=headers)

                                # Process the response
                                if response.status_code == 204:
                                    print(f"transitions sucessfully for row {index}: {response.status_code}")
                                    # requestresponse  = response.json()
                                else:
                                    print(f"transitions failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                                    error = 8
                            except requests.exceptions.RequestException as e:
                                print(f"Error while making transitions Jira API call for row {index}: {e}")   
                                error = 9
                            
                if(error> 0):
                    df.loc[index, 'Completed'] = str(error) + "-Error"
                elif error == 0:
                    df.loc[index, 'Completed'] =  "Done"
            else:
                print("All task time are already logged")
                
        if error  >= 0:        
            book = load_workbook(excel_file_path)
            sheet = book["CreateTask"]
            # updated_values = df['Completed'].tolist()
            # for idx, cell in enumerate(sheet['Completed'], start=3):  # Assuming the 'Age' column is in column B (index 2)
            #     cell.value = updated_values[idx - 2]
            for row_index, row in df.iterrows():
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index + 2, column=col_index, value=value)

            # Save the changes to the Excel file
            book.save(excel_file_path)
            print("Row updated successfully.")
    else:
        print("no new record is present.")
    
    if Logwork ==1:
        try:
            df = pd.read_excel(excel_file_path,sheet_name="Logwork")
        except FileNotFoundError:
            print(f"Error: File '{excel_file_path}' not found.")
            error = -10
            return
        except Exception as e:
            print(f"Error while reading Excel file: {e}")
            error = -11
            return 
        for index, row in df.iterrows():
            error = 0  
            # Extract data from Excel row
            key1        = row['Parentid'].strip()
            comment	    = row['comment']
            started	    = row['started']
            timeSpent   = row['timeSpent']
            Completed   = row['Completed']
            transitions = row['transitions']
            status_name = str()
            callapi     = row['callapi']
            transition_done = row['transition_done']
            print(timeSpent)
            print(key1)
            
            
            if Completed == "New" :
                if callapi ==1:              
                    try:
                        # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                        print(f"{jira_url}/rest/api/2/issue/{key1}")
                        blankpayload= ""
                        response = requests.get(f"{jira_url}/rest/api/2/issue/{key1}",headers=headers)

                        # Process the response
                        if response.status_code == 200:
                            print(f"issue fetched sucessfully for row {index}: {response.json()}")
                            requestresponse  = response.json()
                            status_name = requestresponse["fields"]["parent"]["fields"]["status"]["name"]
                        else:
                            print(f"issue fetched failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                            error = 50
                    except requests.exceptions.RequestException as e:
                        print(f"Error while making issue fetched Jira API call for row {index}: {e}")
                        error = 51
                        
                        
                    print(status_name)    
                    if transitions == 11 and status_name == "Open" :
                            key1= requestresponse["key"]
                            payload_transitions = {
                                "transition": {
                                "id":"11"
                                }
                            }
                            try:
                                # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                                response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/transitions", json=payload_transitions, headers=headers)

                                # Process the response
                                if response.status_code == 204:
                                    print(f"transitions sucessfully for row {index}: {response.status_code}")
                                    # requestresponse  = response.json()
                                else:
                                    print(f"transitions failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                                    error = 52
                            except requests.exceptions.RequestException as e:
                                print(f"Error while making transitions Jira API call for row {index}: {e}")   
                                error = 53
                    
                    payloadlogwork = {
                            
                            "comment": comment,
                            "started": started,
                            "timeSpent":timeSpent
                            
                        }
                    try:
                        # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                        response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/worklog", json=payloadlogwork, headers=headers)

                        # Process the response
                        if response.status_code == 201:
                            print(f"timelogged sucessfully for row {index}: {response.json()}")
                            # requestresponse  = response.json()
                        else:
                            print(f"timelog failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                            error = 54
                    except requests.exceptions.RequestException as e:
                        print(f"Error while making timelog Jira API call for row {index}: {e}")  
                        error = 55 
                    
                    if transition_done == 21 and  response.status_code == 201:
                        # key1= requestresponse["key"]
                        payload_transitions = {
                            "transition": {
                            "id":"21"
                            }
                        }
                        try:
                            # Make the Jira API call to create an issue (you can adjust the API endpoint accordingly)
                            response = requests.post(f"{jira_url}/rest/api/2/issue/{key1}/transitions", json=payload_transitions, headers=headers)

                            # Process the response
                            if response.status_code == 204:
                                print(f"transitions sucessfully for row {index}: {response.status_code}")
                                # requestresponse  = response.json()
                            else:
                                print(f"transitions failed for row {index}. Status code: {response.status_code}, Error: {response.text}")
                                error = 56
                        except requests.exceptions.RequestException as e:
                            print(f"Error while making transitions Jira API call for row {index}: {e}")   
                            error =  57    
                if(error> 0):
                    df.loc[index, 'Completed'] = str(error) + "-Error"
                elif error == 0:
                    df.loc[index, 'Completed'] = str(error) + "Done"
                        
        # df.to_excel(excel_file_path, sheet_name="Logwork", index=False, header=False)
        if(error >= 0):
            book = load_workbook(excel_file_path)
            sheet = book["Logwork"]
            for row_index, row in df.iterrows():
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index + 2, column=col_index, value=value)

            # Save the changes to the Excel file
            book.save(excel_file_path)
            print("Row updated successfully.")
    else:
        print("no new record is present.")

# Example usage:
if __name__ == "__main__":
    excel_file_path = "DailyTask.xlsx"
    jira_url = "https://corecard.atlassian.net"
    username = "prasoon.parashar@corecard.com"
    api_token = "ATATT3xFfGF07ucHHFlug1ho0U4Y8lZ6oaQba5S8u1aNn-d6gtsX5cBz5ytLHOxf-Gvw3BMv4S9GwiSHml8u5Pgvr3-Ln8TBNS8I0UhyKUB6wxfXlPrD8q-puIfKZxEJp5Dnc8QpENbEfGyhg_G9yw3rDMbS5Y_g-RrfNGKukOiSkQ2PnPUUvR8=5FC8B27B"
    projectid = 10028
    jirauserid= "5fc95d49d670b8006e1ae556" 
    componentid= 10083
    read_excel_and_call_jira_api(excel_file_path, jira_url, username, api_token)
